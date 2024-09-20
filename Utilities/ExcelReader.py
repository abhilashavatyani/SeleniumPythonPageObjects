import openpyxl

# open load workbook. provide the path. will be providing dynamic path.
workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")

# define the sheet from which we need to read the data
sheet = workbook["NewCarsTest"]

# check the row count and total column count
totalrows = sheet.max_row
totalcols = sheet.max_column

print("total rows:", str(totalrows), "total cols are: ", str(totalcols))

# read the first row and first column
print(sheet.cell(row=2, column=1).value)

# print entire data
for rows in range(1, totalrows + 1):
    for cols in range(1, totalcols + 1):
        print(sheet.cell(row=rows, column=cols).value, end="  ")
    print()
