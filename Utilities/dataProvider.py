import openpyxl


def get_data(sheetName):
    # 101. data driven test through excel.
    workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column

    # parametrized the data
    mainList = []

    for r in range(2, totalrows + 1):
        dataList = []
        for c in range(1, totalcols + 1):
            data = sheet.cell(row=r, column=c).value
            dataList.insert(c, data)
        mainList.insert(r, dataList)
    print(mainList)
    return mainList
