# this test will have parameterized marks and data. get_data() dynamic data from excel
# will use allure part as well
# will have fixtures. will shift fixtures to "conftest.py". most of the fixtures will be keeping here.


import pytest
import openpyxl
from Pages.RegistrationPage import Registration
from TestCases.BaseTest import BaseTest
from Utilities import dataProvider

import logging
from Utilities.LogUtil import Logger
# "__name__" to get the current file name
log = Logger(__name__, logging.INFO)

## move the below code to dataProvider.py
# def get_data():
#     # 101. data driven test through excel.
#     workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
#     sheet = workbook["LoginTest"]
#     totalrows = sheet.max_row
#     totalcols = sheet.max_column
#
#     # parametrized the data
#     mainList = []
#
#     for r in range(2, totalrows + 1):
#         dataList = []
#         for c in range(1, totalcols + 1):
#             data = sheet.cell(row=r, column=c).value
#             dataList.insert(c, data)
#         mainList.insert(r, dataList)
#     print(mainList)
#     return mainList

class Test_SignUp(BaseTest):

    #@pytest.mark.usefixtures("log_on_failure")  # 82. capturing screenshot in case of failure      ##moved all the usefixtures to baseTest.py
    @pytest.mark.parametrize("name, phoneNum, email, country, city, username, password",
                             dataProvider.get_data("LoginTest"))
    def test_doSignUp(self, name, phoneNum, email, country, city, username, password):
                      #, get_browser):                                                #moved all the usefixtures to baseTest.py
        #self.driver = get_browser

        #log
        log.logger.info("Test Do_SignUp started.")
        # to execute the test we need to create object of RegistrationPage
        regPage = Registration(self.driver)  # create the object
        regPage.fillform(name, phoneNum, email, country, city, username, password)
        log.logger.info("Test Do_SignUp successfully executed.")